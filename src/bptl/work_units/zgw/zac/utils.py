import io
from datetime import date, datetime, timedelta, timezone
from typing import Any, Dict, Iterable, List, Optional, Tuple

import pytz
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from zgw_consumers.api_models.constants import RolTypes

from bptl.camunda.constants import AssigneeTypeChoices
from bptl.tasks.models import BaseTask

from .client import get_client

# -------------------------
# Shared helpers
# -------------------------


def _parse_dt_any(val: Any) -> Optional[datetime]:
    """Parse a value into datetime if possible; supports date, ISO (incl. 'Z'), and common formats."""
    if isinstance(val, datetime):
        return val
    if isinstance(val, date):
        return datetime(val.year, val.month, val.day)
    if isinstance(val, str) and val:
        s = val.replace("Z", "+00:00")  # support trailing 'Z' (UTC)
        try:
            return datetime.fromisoformat(s)  # handles offsets and naive ISO
        except ValueError:
            pass
        for fmt in ("%Y-%m-%d", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M:%S"):
            try:
                return datetime.strptime(val, fmt)
            except ValueError:
                continue
    return None


def _to_excel_date(val: Any) -> Optional[date]:
    """
    Parse various date/datetime strings (incl. trailing 'Z' or offsets) and
    return a timezone-naive date object for Excel.
    """
    dt = _parse_dt_any(val)
    if not dt:
        return None
    if dt.tzinfo is not None:
        dt = dt.astimezone(timezone.utc).replace(tzinfo=None)
    return dt.date()


def _load_wb(binary: bytes) -> Workbook:
    return load_workbook(io.BytesIO(binary))


def _wb_to_bytes(wb: Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _ensure_new_sheet(wb: Workbook, name: str) -> Worksheet:
    """Create a fresh sheet with given name (remove existing if present)."""
    if name in wb.sheetnames:
        wb.remove(wb[name])
    return wb.create_sheet(title=name)


def _daterange(start: datetime, end: datetime) -> Iterable[datetime]:
    """Inclusive date range at day resolution."""
    cur = start
    one_day = timedelta(days=1)
    while cur.date() <= end.date():
        yield cur
        cur = cur + one_day


def _bold_and_freeze_header(ws: Worksheet) -> None:
    """Bold the first row and freeze panes below it."""
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"


def _apply_autofilter(ws: Worksheet, num_cols: int) -> None:
    """Apply an AutoFilter over A1:<last_col_letter><last_row>."""
    last_row = ws.max_row
    last_col_letter = get_column_letter(num_cols)
    ws.auto_filter.ref = f"A1:{last_col_letter}{last_row}"


def _autosize_columns(ws: Worksheet, num_cols: int, padding: int = 2) -> None:
    """Autosize columns to max content width (including header)."""
    for col_idx in range(1, num_cols + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for cell in ws[col_letter]:
            val = "" if cell.value is None else str(cell.value)
            cell_len = max((len(line) for line in val.splitlines()), default=0)
            if cell_len > max_len:
                max_len = cell_len
        ws.column_dimensions[col_letter].width = max_len + padding


# -------------------------
# Domain helpers
# -------------------------


def get_betrokkene_identificatie(rol: Dict[str, Any], task: BaseTask) -> Dict[str, Any]:
    betrokkene_identificatie: Dict[str, Any] = rol.get("betrokkeneIdentificatie") or {}

    if rol.get(
        "betrokkeneType"
    ) == RolTypes.medewerker and f"{AssigneeTypeChoices.user}:" in betrokkene_identificatie.get(
        "identificatie", ""
    ):
        # Only fetch details if initials/prefix/surname are all missing
        if not (
            betrokkene_identificatie.get("voorletters")
            or betrokkene_identificatie.get("voorvoegsel_achternaam")
            or betrokkene_identificatie.get("achternaam")
        ):
            with get_client(task) as client:
                betrokkene_identificatie = client.post(
                    "api/core/rollen/medewerker/betrokkeneIdentificatie",
                    json={"betrokkeneIdentificatie": betrokkene_identificatie},
                ).get("betrokkeneIdentificatie", {})

    return betrokkene_identificatie


def get_last_month_period(
    timezone_str: str = "Europe/Amsterdam",
) -> Tuple[datetime, datetime]:
    """
    Returns two timezone-aware ISO datetime strings:
    1) First day of previous month 00:00:00
    2) Last day of previous month 23:59:59
    """
    tz = pytz.timezone(timezone_str)
    today = datetime.now(tz)
    first_of_this_month = today.replace(
        day=1, hour=0, minute=0, second=0, microsecond=0
    )
    last_month_end = first_of_this_month - timedelta(seconds=1)
    last_month_start_iso = last_month_end.replace(
        day=1, hour=0, minute=0, second=0, microsecond=0
    )
    last_month_end_iso = last_month_end.replace(
        hour=23, minute=59, second=59, microsecond=0
    )
    return last_month_start_iso, last_month_end_iso


# -------------------------
# Report: Zaken
# -------------------------


def create_zaken_report_xlsx(results: List[Dict[str, Any]]) -> bytes:
    """
    Create an Excel workbook for ZAAK results with:
    - Columns: Identificatie, Omschrijving, Zaaktype, Registratiedatum, Initiator, Object, Objecttype, Aantal Informatieobjecten
    - One row per *object* in `objecten` (list of {"object","objecttype"}). If none, a single row with blanks.
    - Sorted by registratiedatum (oldest first, blanks last)
    - registratiedatum written as Excel date cells (yyyy-mm-dd)
    - Bold + frozen header row, AutoFilter, auto-sized columns
    """
    # Display headers (explicit to avoid accidentally including "objecten")
    headers_display: List[str] = [
        "Identificatie",
        "Omschrijving",
        "Zaaktype",
        "Registratiedatum",
        "Initiator",
        "Object",
        "Objecttype",
        "Aantal Informatieobjecten",
    ]

    def _normalize_keys(row: Dict[str, Any]) -> Dict[str, Any]:
        """Normalize known camelCase → snake_case keys."""
        r = row.copy()
        if "aantalInformatieobjecten" in r and "aantal_informatieobjecten" not in r:
            r["aantal_informatieobjecten"] = r["aantalInformatieobjecten"]
        return r

    # Normalize keys first
    normalized = [_normalize_keys(r) for r in (results or [])]

    # Sort by registratiedatum using timezone-naive date for sorting
    sorted_rows = sorted(
        normalized,
        key=lambda item: (
            (d := _to_excel_date(item.get("registratiedatum"))) is None,
            d or date.max,
        ),
    )

    # Build workbook
    wb: Workbook = Workbook()
    ws: Worksheet = wb.active
    ws.title = "Zaken"
    ws.append(headers_display)
    _bold_and_freeze_header(ws)

    # Helper to explode objecten to list[{"object","objecttype"}]
    def _explode_objects(obj_field: Any) -> List[Dict[str, str]]:
        """
        Expected shape now: list of {"object": <str>, "objecttype": <str>}
        Also supports legacy: string or list of scalars.
        """
        if isinstance(obj_field, list):
            out: List[Dict[str, str]] = []
            for it in obj_field:
                if isinstance(it, dict):
                    out.append(
                        {
                            "object": str(it.get("object", "") or ""),
                            "objecttype": str(it.get("objecttype", "") or ""),
                        }
                    )
                else:
                    # list of scalars -> treat as single 'object' values
                    out.append({"object": str(it or ""), "objecttype": ""})
            return out
        if isinstance(obj_field, str):
            return [{"object": obj_field, "objecttype": ""}]
        return []

    # Write rows (explode per object) — NEVER append the raw "objecten" list
    for row in sorted_rows:
        excel_date = _to_excel_date(row.get("registratiedatum"))

        objects_list = _explode_objects(row.get("objecten"))
        if not objects_list:
            # still write a single row with empty object/objecttype
            objects_list = [{"object": "", "objecttype": ""}]

        for obj_entry in objects_list:
            out: List[Any] = [
                row.get("identificatie", ""),
                row.get("omschrijving", ""),
                row.get("zaaktype", ""),
                None,  # placeholder for date cell
                row.get("initiator", ""),
                obj_entry.get("object", ""),
                obj_entry.get("objecttype", ""),
                row.get("aantal_informatieobjecten", 0) or 0,
            ]
            ws.append(out)

            # Format registratiedatum cell as *date* (naive) if possible
            if excel_date:
                reg_col_idx = 4  # "Registratiedatum" column
                cell = ws.cell(row=ws.max_row, column=reg_col_idx)
                cell.value = excel_date
                cell.number_format = "yyyy-mm-dd"

    # Finish
    _apply_autofilter(ws, len(headers_display))
    _autosize_columns(ws, len(headers_display))
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


# -------------------------
# Report: Users
# -------------------------


def add_users_sheet_xlsx(
    report_excel: bytes,
    results_user_logins: List[Dict[str, Any]],
    start_period: Optional[str] = None,
    end_period: Optional[str] = None,
    date_format: str = "%Y-%m-%d",
    sheet_name: str = "Gebruikers",
) -> bytes:
    """
    Add a 'Gebruikers' sheet with columns:
        Naam, Email, Gebruikersnaam, Totaal, <each date between start_period and end_period>
    Rows are ordered by 'Naam'. If a user's loginsPerDay lacks a date, fill 0.
    """
    wb: Workbook = _load_wb(report_excel)

    # Date range for columns (supports ISO, 'Z', etc.)
    if start_period and end_period:
        start_dt = _parse_dt_any(start_period) or datetime.fromisoformat(start_period)
        end_dt = _parse_dt_any(end_period) or datetime.fromisoformat(end_period)
        if end_dt < start_dt:  # defensive swap
            start_dt, end_dt = end_dt, start_dt
    else:
        min_d: Optional[datetime] = None
        max_d: Optional[datetime] = None
        for u in results_user_logins or []:
            per_day = u.get("loginsPerDay", {}) or {}
            for ds in per_day.keys():
                d = _parse_dt_any(ds) or datetime.strptime(str(ds), "%Y-%m-%d")
                if min_d is None or d < min_d:
                    min_d = d
                if max_d is None or d > max_d:
                    max_d = d
        if min_d is None or max_d is None:
            today = datetime.today().replace(hour=0, minute=0, second=0, microsecond=0)
            min_d = max_d = today
        start_dt, end_dt = min_d, max_d

    date_headers = [d.strftime(date_format) for d in _daterange(start_dt, end_dt)]

    base_headers = ["naam", "email", "gebruikersnaam", "totaal"]
    formatted_base_headers = [h.replace("_", " ").title() for h in base_headers]
    headers = [*formatted_base_headers, *date_headers]

    ws = _ensure_new_sheet(wb, sheet_name)
    ws.append(headers)
    _bold_and_freeze_header(ws)

    # Sort by name
    sorted_users = sorted(
        results_user_logins or [], key=lambda u: (u.get("naam") or "").lower()
    )

    # Write rows
    for u in sorted_users:
        naam = u.get("naam", "") or ""
        email = u.get("email", "") or ""
        gebruikersnaam = u.get("gebruikersnaam", "") or ""
        totaal = int(u.get("totalLogins", 0) or 0)

        per_day: Dict[str, Any] = u.get("loginsPerDay", {}) or {}
        normalized_per_day: Dict[str, int] = {}
        for k, v in per_day.items():
            dt = _parse_dt_any(k) or datetime.strptime(str(k), "%Y-%m-%d")
            normalized_per_day[dt.strftime(date_format)] = int(v or 0)

        row = [naam, email, gebruikersnaam, totaal]
        row.extend(normalized_per_day.get(dh, 0) for dh in date_headers)
        ws.append(row)

    _apply_autofilter(ws, len(headers))
    _autosize_columns(ws, len(headers))
    return _wb_to_bytes(wb)


# -------------------------
# Report: Informatieobjecten
# -------------------------


def add_informatieobjecten_sheet_xlsx(
    report_excel: bytes,
    results_informatieobjecten: List[Dict[str, Any]],
    sheet_name: str = "Informatieobjecten",
    date_out_format: str = "%Y-%m-%d %H:%M:%S",  # kept for signature compatibility (not used)
) -> bytes:
    """
    Add an 'Informatieobjecten' sheet with columns:
        Auteur, Bestandsnaam, Beschrijving, Informatieobjecttype, Creatiedatum,
        Gerelateerde Zaken, Zaaktype

    Behavior:
    - Each related zaak in 'gerelateerdeZaken' becomes its own row (exploded).
    - 'Zaaktype' column contains only the zaaktype.omschrijving (substring after ':').
    - Sorting: creatiedatum (asc, None last) → informatieobjecttype → bestandsnaam → auteur → zaaktype
    - 'creatiedatum' is written as an Excel date (yyyy-mm-dd) when parseable; otherwise the raw string is written.
    - Header bold + frozen, AutoFilter on, columns auto-sized.
    """
    wb = _load_wb(report_excel)
    ws = _ensure_new_sheet(wb, sheet_name)

    headers = [
        "Auteur",
        "Bestandsnaam",
        "Beschrijving",
        "Informatieobjecttype",
        "Creatiedatum",
        "Gerelateerde Zaken",
        "Zaaktype",
    ]
    ws.append(headers)
    _bold_and_freeze_header(ws)

    def _extract_zaaktype(s: Any) -> str:
        """Extract the part after the first colon (':') — the zaaktype.omschrijving."""
        if not s:
            return ""
        txt = str(s)
        return txt.split(":", 1)[1].strip() if ":" in txt else txt.strip()

    # Build exploded rows, carrying excel_date (date) and raw creatiedatum (str/Any) for fallback
    exploded_rows: List[Tuple[Tuple, List[Any], Optional[date], Any]] = []

    for it in results_informatieobjecten or []:
        auteur = it.get("auteur") or ""
        bestandsnaam = it.get("bestandsnaam") or ""
        beschrijving = it.get("beschrijving") or ""
        iot = it.get("informatieobjecttype") or ""

        dt_raw = it.get("creatiedatum")
        # Parse to naive date for Excel using your helper if available
        try:
            excel_date = _to_excel_date(dt_raw)  # Optional[date]
        except NameError:
            excel_date = None

        related = it.get("gerelateerdeZaken")

        if not related:
            row_vals = [auteur, bestandsnaam, beschrijving, iot, None, "", ""]
            sort_key = (
                excel_date is None,
                excel_date or date.max,
                iot.lower(),
                bestandsnaam.lower(),
                auteur.lower(),
                "",
            )
            exploded_rows.append((sort_key, row_vals, excel_date, dt_raw))
        else:
            for entry in related:
                full_rel = str(entry) if entry is not None else ""
                zaaktype_only = _extract_zaaktype(full_rel)
                row_vals = [
                    auteur,
                    bestandsnaam,
                    beschrijving,
                    iot,
                    None,  # we'll set date cell or raw string after append
                    full_rel,
                    zaaktype_only,
                ]
                sort_key = (
                    excel_date is None,
                    excel_date or date.max,
                    iot.lower(),
                    bestandsnaam.lower(),
                    auteur.lower(),
                    zaaktype_only.lower(),
                )
                exploded_rows.append((sort_key, row_vals, excel_date, dt_raw))

    # Sort and write (set Excel date if available; else write raw string if present)
    for sort_key, row_vals, excel_date, dt_raw in sorted(
        exploded_rows, key=lambda x: x[0]
    ):
        ws.append(row_vals)
        cell = ws.cell(row=ws.max_row, column=5)  # "Creatiedatum" column
        if excel_date:
            cell.value = excel_date
            cell.number_format = "yyyy-mm-dd"
        elif isinstance(dt_raw, str) and dt_raw:
            cell.value = dt_raw  # fallback to raw string

    _apply_autofilter(ws, len(headers))
    _autosize_columns(ws, len(headers))
    return _wb_to_bytes(wb)
